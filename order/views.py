from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from address.models import Address
from cart.models import Cart
from cart.models import CartItems
from decimal import Decimal
from order.models import Order, OrderItem
from django.shortcuts import get_object_or_404
from pincode.models import Pincode
from django.conf import settings
import razorpay # type: ignore
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from wallet.models import Wallet
from django.utils import timezone
from coupon.models import Coupon
import json
from coupon.models import UserCoupon
from wallet.models import WalletTransaction
from offer.models import Product_Offers, Category_Offers
from django.db.models import Sum, F, Q, Count, Value, DecimalField, IntegerField
from datetime import datetime, date, timedelta
from django.contrib import messages
from django.db.models.functions import TruncDate, Coalesce
from django.db.models import Value
from reportlab.lib.pagesizes import letter # type: ignore
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle # type: ignore
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle # type: ignore
from reportlab.lib import colors # type: ignore
from reportlab.lib.units import inch # type: ignore
import io
from openpyxl import Workbook # type: ignore
from openpyxl.styles import Font, PatternFill # type: ignore
from xhtml2pdf import pisa # type: ignore
import random
from io import BytesIO
from order.models import Invoice
from django.views.decorators.cache import never_cache
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import ExpressionWrapper, F, DecimalField, Case, When


# Create your views here.

razorpay_client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))


@login_required(login_url='login')
@never_cache
def place_order(request):
    user = request.user
    address = Address.objects.filter(user=user, is_listed=True)
    cart = Cart.objects.get(user=user)
    cart_items = CartItems.objects.filter(cart=cart, products__is_listed = True, products__category__is_listed = True)
    coupon_code = request.session.get('coupon_code')
    discount_value = Decimal(request.session.get('discount_value', '0.00'))
    coupon = Coupon.objects.all()
    now = timezone.now()

    subtotal_price = Decimal('0.00')
    cart_items_with_subtotals = []

    total_offer_discount = Decimal('0.00')
    for item in cart_items:
        quantity = Decimal(item.quantity)
        product = item.products
        actual_price = Decimal(product.price)

        product_offer = Product_Offers.objects.filter(
            product=product, valid_from__lte=now, valid_to__gte=now
        ).first()
        category_offer = Category_Offers.objects.filter(
            category=product.category, valid_from__lte=now, valid_to__gte=now
        ).first()

        product_discount = actual_price * (Decimal(product_offer.offer_percentage) / 100) if product_offer else Decimal('0.00')
        category_discount = actual_price * (Decimal(category_offer.offer_percentage) / 100) if category_offer else Decimal('0.00')
        best_offer_discount = max(product_discount, category_discount)
        
        if best_offer_discount:
            total_offer_discount += best_offer_discount * quantity

        offer_price = actual_price - best_offer_discount
        item_subtotal = offer_price * quantity
        subtotal_price = round(subtotal_price + item_subtotal, 2)

        cart_items_with_subtotals.append({
            'item': item,
            'offer_price': offer_price,
            'item_subtotal': item_subtotal,
        })
    if coupon_code:
        applied_coupon = Coupon.objects.filter(code=coupon_code).first()
        if applied_coupon and subtotal_price < applied_coupon.minimum_purchase_amount:
            request.session.pop('coupon_code', None)
            request.session.pop('discount_value', None)
            discount_value = Decimal('0.00')
            messages.warning(request, f"Coupon '{coupon_code}' has been removed as the subtotal is below the minimum purchase requirement.")

    
    delivery_charge = Decimal('0.00') if subtotal_price > Decimal('500.00') else Decimal('40.00')
    total_price = subtotal_price - discount_value + delivery_charge

    available_coupons = []
    active_coupons = Coupon.objects.filter(active=True)
    
    for coupon in active_coupons:
        # Check if user has already used this coupon
        user_coupon_exists = UserCoupon.objects.filter(
            user=user, 
            coupon=coupon,
            used=True
        ).exists()
        
        if (not user_coupon_exists and 
            coupon.minimum_purchase_amount <= subtotal_price and 
            coupon.valid_from <= now.date() <= coupon.valid_upto):
            available_coupons.append(coupon)


    if request.method == 'POST':
        selected_address_id = request.POST.get('address')
        payment_type = request.POST.get('option')

        if not selected_address_id or not payment_type:
            return render(request, 'user/checkout.html', {
                'address': address,
                'cart_items': cart_items_with_subtotals,
                'cart': cart,
                'subtotal_price': round(subtotal_price, 2),
                'delivery_charge': round(delivery_charge, 2),
                'total_price': round(total_price, 2),
                'error_message': 'Please select an address or payment type.'
            })

        selected_address = Address.objects.get(id=selected_address_id)

        if payment_type == 'Razor pay':
            request.session['selected_address'] = selected_address.id
            razorpay_order = razorpay_client.order.create({
                'amount': int(total_price * Decimal('100')),
                'currency': 'INR',
                'payment_capture': '1'
            })

            return render(request, 'user/razorpay.html', {
                'razorpay_order_id': razorpay_order['id'],
                'razorpay_key_id': settings.RAZORPAY_KEY_ID,
                'amount': total_price,
                'selected_address': selected_address,
            })
        
        if payment_type == 'Wallet':
            wallet = Wallet.objects.get(user=user)
            if wallet.balance < total_price:
                return render(request, 'user/checkout.html',{
                    'address':address,
                    'cart_items':cart_items_with_subtotals,
                    'cart':cart,
                    'subtotal_price': round(subtotal_price, 2),
                    'delivery_charge': round(delivery_charge, 2),
                    'total_price': round(total_price, 2),
                    'error_message':'Insufficient balance in your wallet. Please choose another payment method.'
                })
            wallet.balance -= total_price
            wallet.save()

            WalletTransaction.objects.create(
                wallet = wallet,
                transaction_type = 'Debited',
                amount = total_price,
            )
        
            payment_status = 'Success'

            new_order=Order.objects.create(
                user=user,
                address=selected_address,
                payment_type=payment_type,
                payment_status=payment_status,
                total_price=total_price,
                coupon_code=coupon_code,
                coupon_discount=discount_value,
                offer_discount=total_offer_discount
            )

            item_discount = Decimal('0.00')
            if discount_value > 0:
                if len(cart_items) > 1:
                    item_discount = discount_value / len(cart_items)
                else:
                    item_discount = discount_value

            cart_items_with_subtotals = []

            for item in cart_items:
                quantity = Decimal(item.quantity)
                item_price = Decimal(item.products.price)
                item_subtotal_price = (item_price * quantity) - item_discount
                cart_items_with_subtotals.append({
                    'item': item,
                    'item_subtotal': item_subtotal_price
                })

                OrderItem.objects.create(
                    order=new_order,
                    product=item.products,
                    quantity=quantity,
                    price=item_price,
                    subtotal_price=item_subtotal_price  # Store item subtotal after applying discount
                )

                if payment_status== 'Success':
                    item.products.stock -= quantity
                    item.products.save()

            cart_items.delete()

            if coupon_code:
                user_coupon = UserCoupon.objects.get(user=user, coupon__code=coupon_code)
                user_coupon.used = True
                user_coupon.save()


            order_items = OrderItem.objects.filter(order=new_order)

            email_subject = 'order confirmation'
            email_body = render_to_string('user/order_confirming_email.html', {
                'user': user,
                'order': new_order,
                'order_items': order_items
            })
            email_message = EmailMessage(email_subject, email_body, to=[user.email])
            email_message.content_subtype = "html"
            email_message.send()

            request.session.pop('coupon_code', None)
            request.session.pop('discount_value', None)

            return redirect('order_success',)

        if payment_type == 'Cash on Delivery' and total_price > Decimal('1000.00'):
            return render(request, 'user/checkout.html', {
                'address': address,
                'cart_items': cart_items_with_subtotals,
                'cart': cart,
                'subtotal_price': round(subtotal_price, 2),
                'delivery_charge': round(delivery_charge, 2),
                'total_price': round(total_price, 2),
                'error_message': 'Cash on Delivery is not available for orders above ₹1000.'
            })

        payment_status = 'Success'

        # Create a new order with the discounted price and applied coupon
        new_order = Order.objects.create(
            user=request.user,
            address=selected_address,
            total_price=total_price,
            payment_type=payment_type,
            coupon_code=coupon_code,
            coupon_discount=discount_value,
            offer_discount=total_offer_discount
        )

        item_discount = Decimal('0.00')
        if discount_value > 0:
            if len(cart_items) > 1:
                item_discount = discount_value / len(cart_items)
            else:
                item_discount = discount_value

        cart_items_with_subtotals = []

        for item in cart_items:
            quantity = Decimal(item.quantity)
            item_price = Decimal(item.products.price)
            item_subtotal_price = (item_price * quantity) - item_discount

            cart_items_with_subtotals.append({
                'item': item,
                'item_subtotal': item_subtotal_price
            })

            OrderItem.objects.create(
                order=new_order,
                product=item.products,
                quantity=quantity,
                price=item_price,
                subtotal_price=item_subtotal_price
            )

            if payment_status == 'Success':
                item.products.stock -= quantity
                item.products.save()

        cart_items.delete()

        if coupon_code:
            user_coupon = UserCoupon.objects.get(user=user, coupon__code=coupon_code)
            user_coupon.used = True
            user_coupon.save()

        # Clear coupon data from the session after order placement
        request.session.pop('coupon_code', None) 
        request.session.pop('discount_value', None)

        return render(request, 'user/order_success.html')

    return render(request, 'user/checkout.html', {
        'address': address,
        'cart_items': cart_items_with_subtotals,
        'cart': cart,
        'subtotal_price': round(subtotal_price, 2),
        'delivery_charge': round(delivery_charge, 2),
        'total_price': round(total_price, 2),
        'coupon': available_coupons
    })

@login_required(login_url='login')
@never_cache
def add_address_in_checkout(request):
    
     
    if request.method == 'POST':
 
        name  =  request.POST.get("name")
        phone_number= request.POST.get("phone")
        alternative_phone = request.POST.get("alt_phone")
        pincode_value = request.POST.get("pincode")
        locality = request.POST.get("locality")
        landmark = request.POST.get("landmark")
        district = request.POST.get("district")
        state = request.POST.get("state")
        country = request.POST.get("country")
        address = request.POST.get("address")
        address_type = request.POST.get("addressType")
        pincode_instance = get_object_or_404(Pincode, pincode=pincode_value)

        Address.objects.create(
            user = request.user,
            full_name =name,
            phone_number = phone_number,
            alternative_phone_number =  alternative_phone,
            pincode = pincode_instance,
            locality =  locality,
            landmark = landmark,
            district =  district,
            state = state,
            country = country,
            address = address,
            address_type = address_type
        )
        return redirect('checkout')

    return render(request, "user/checkout_add_address.html")

@login_required(login_url='login')
@never_cache
def user_order(request):
    user = request.user
    orders = Order.objects.filter(user=user).order_by('-created_at')
    paginator = Paginator(orders, 10)
    page = request.GET.get('page', 1)
    
    try:
        paginated_orders = paginator.page(page)
    except PageNotAnInteger:
        paginated_orders = paginator.page(1)
    except EmptyPage:
        paginated_orders = paginator.page(paginator.num_pages)

    return render(request, 'user/user_order.html', {'orders': paginated_orders})

@login_required(login_url='login')
@never_cache
def user_single_order_items(request, order_id):

    order = get_object_or_404(Order, id = order_id)
    order_items = OrderItem.objects.filter(order=order)

    for item in order_items:
        print("Item:",item.subtotal_price )
        print(len(order_items))
    
    context = {
        'order': order,
        'order_items':order_items,
    }

    return render(request, 'user/user_single_order_items.html', context)

@login_required(login_url='login')
@never_cache
def user_order_details(request,item_id):
    order_items = get_object_or_404(OrderItem, id=item_id)
    order = order_items.order

    original_price = order_items.quantity * order_items.price
    item_coupon_discount = original_price - order_items.subtotal_price

    context = {
       
        'order_items': order_items,
        'order' : order,
        'item_coupon_discount': item_coupon_discount,
    }
    return render(request,"user/user_order_details.html",context)

def user_singleitem_cancel(request,order_item_id):
    order_item =  get_object_or_404(OrderItem,id=order_item_id)

    if order_item.status in ["Order Pending", "Order confirmed","Shipped"]:
        order_item.status = "Cancelled"
        order_item.save()

        process_refund(order_item)
        order_item.product.stock += order_item.quantity
        order_item.product.save()

    return redirect('user_single_order_items', order_id = order_item.order.id)

@login_required(login_url='admin_login')
@never_cache
def admin_order_list(request):
    orders = Order.objects.all().order_by('-created_at')
    paginator = Paginator(orders, 10)
    page = request.GET.get('page', 1)
    
    try:
        paginated_orders = paginator.page(page)
    except PageNotAnInteger:
        paginated_orders = paginator.page(1)
    except EmptyPage:
        paginated_orders = paginator.page(paginator.num_pages)

    return render(request, 'admin_order_list.html', {'orders': paginated_orders})

@login_required(login_url='admin_login')
@never_cache
def admin_single_order_details(request, id):
    order = get_object_or_404(Order,id=id)
    order_items = order.items.all()

    context = {
        'order': order,
        'order_items': order_items
    }
    return render(request, 'admin_single_order_details.html', context)

def update_order_status(request, id):
    order = get_object_or_404(Order, id=id)
    
    if request.method == 'POST':
        for item in order.items.all():
            new_status = request.POST.get(f'status_{item.id}')
            old_status = item.status
            if new_status:
                if old_status not in ['Cancelled', 'Approve Return'] and new_status in ['Cancelled', 'Approve Return']:
                    process_refund(item)
                item.status = new_status
                item.save()
            
                if new_status == "Cancelled" and old_status != "Cancelled":
                    item.product.stock += item.quantity
                    item.product.save()
        
        return redirect('admin_order_list')
    
    return redirect('admin_single_order_details', id=id)

@csrf_exempt
def razorpay_payment(request):
    if request.method == 'POST':
        payment_id = request.POST.get('razorpay_payment_id')
        razorpay_order_id = request.POST.get('razorpay_order_id')
        razorpay_signature = request.POST.get('razorpay_signature')

        user = request.user
        cart = Cart.objects.get(user=user)
        cart_items = CartItems.objects.filter(cart=cart)
        selected_address_id = request.session.get('selected_address')
        selected_address = Address.objects.get(id=selected_address_id)
        subtotal_price = Decimal('0.00')
        cart_items_with_subtotals = []
        now = timezone.now()

        total_offer_discount = Decimal('0.00')

        for item in cart_items:
            quantity = Decimal(item.quantity)
            product = item.products
            actual_price = Decimal(product.price)

            product_offer = Product_Offers.objects.filter(
                product=product, valid_from__lte=now, valid_to__gte=now
            ).first()
            category_offer = Category_Offers.objects.filter(
                category=product.category, valid_from__lte=now, valid_to__gte=now
            ).first()

            product_discount = actual_price * (Decimal(product_offer.offer_percentage) / 100) if product_offer else Decimal('0.00')
            category_discount = actual_price * (Decimal(category_offer.offer_percentage) / 100) if category_offer else Decimal('0.00')
            best_offer_discount = max(product_discount, category_discount)

            if best_offer_discount:
                total_offer_discount += best_offer_discount * quantity

            offer_price = actual_price - best_offer_discount
            item_subtotal = offer_price * quantity
            subtotal_price += item_subtotal

            cart_items_with_subtotals.append({
                'item': item,
                'offer_price': offer_price,
                'item_subtotal': item_subtotal
            })


        delivery_charge = Decimal('0.00') if subtotal_price > Decimal('500.00') else Decimal('40.00')

        discount_value = Decimal(request.session.get('discount_value', '0.00'))
        coupon_code = request.session.get('coupon_code')

        if coupon_code:
            try:
                applied_coupon = Coupon.objects.get(code=coupon_code, active=True)
                user_coupon, created = UserCoupon.objects.get_or_create(user=user, coupon=applied_coupon)
                

                if user_coupon.used:
                    return JsonResponse({"error": "You have already used this coupon."})

                if subtotal_price >= applied_coupon.minimum_purchase_amount:
                    discount_value = applied_coupon.discount_value
                    subtotal_price -= discount_value
                else:
                    discount_value = Decimal('0.00')
            except Coupon.DoesNotExist:
                discount_value = Decimal('0.00')


        total_price = subtotal_price + delivery_charge

        payment_status = 'Failure'
        
        try:
            razorpay_client.utility.verify_payment_signature({
                'razorpay_order_id': razorpay_order_id,
                'razorpay_payment_id': payment_id,
                'razorpay_signature': razorpay_signature
            })
            payment_status = 'Success'
        except razorpay.errors.SignatureVerificationError:
            pass

        new_order = Order.objects.create(
            user=user,
            address=selected_address,
            payment_type='Razor pay',
            total_price=total_price,
            payment_status=payment_status,
            coupon_code=coupon_code,
            coupon_discount=discount_value,
            offer_discount=total_offer_discount
        )

        item_discount = Decimal('0.00')
        if discount_value > 0:
            if len(cart_items) > 1:
                item_discount = discount_value / len(cart_items)
            else:
                item_discount = discount_value

        for item in cart_items:
            quantity = Decimal(item.quantity)
            actual_price = Decimal(item.products.price)

            product_offer = Product_Offers.objects.filter(
                product=item.products, valid_from__lte=now, valid_to__gte=now
            ).first()
            category_offer = Category_Offers.objects.filter(
                category=item.products.category, valid_from__lte=now, valid_to__gte=now
            ).first()

            product_discount = actual_price * (Decimal(product_offer.offer_percentage) / 100) if product_offer else Decimal('0.00')
            category_discount = actual_price * (Decimal(category_offer.offer_percentage) / 100) if category_offer else Decimal('0.00')
            best_offer_discount = max(product_discount, category_discount)

            offer_price = actual_price - best_offer_discount
            item_subtotal_price = (offer_price * quantity) - item_discount

            OrderItem.objects.create(
                order=new_order,
                product=item.products,
                quantity=item.quantity,
                price=offer_price,
                subtotal_price=item_subtotal_price
            )

        cart_items.delete()

        if payment_status == 'Success':
            for item in cart_items:
                item.products.stock -= Decimal(item.quantity)
                item.products.save()


            if coupon_code:
                user_coupon.used = True
                user_coupon.save()

            order_items = OrderItem.objects.filter(order=new_order)

            email_subject = 'order confirmation'
            email_body = render_to_string('user/order_confirming_email.html', {
                'user': user,
                'order': new_order,
                'order_items': order_items
            })
            email_message = EmailMessage(email_subject, email_body, to=[user.email])
            email_message.content_subtype = "html"
            email_message.send()

            request.session.pop('coupon_code', None)
            request.session.pop('discount_value', None)

            return redirect('order_success',)
        else:
            return render(request, 'user/order_success.html', {'error_message': 'Payment failed. Your order has been placed but payment was not completed.'})

    return render(request, 'user/order_success.html', {'error_message': 'Invalid request.'})


def retry_payment(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    total_price = order.total_price

    razorpay_order = razorpay_client.order.create({
        "amount": int(total_price * 100),  
        "currency": "INR",
        "payment_capture": "1"  
    })
    
    
    context = {
        "order": order,
        "razorpay_order_id": razorpay_order['id'],
        "razorpay_key": settings.RAZORPAY_KEY_ID,  
        "amount": order.total_price * 100, 
    }
    
    return render(request, 'user/razorpay_retry_payment.html', context)

@csrf_exempt
def handle_payment(request):
    if request.method == "POST":
        payment_id = request.POST.get('razorpay_payment_id')
        order_number = request.POST.get('razorpay_order_id')  
        signature = request.POST.get('razorpay_signature')
        db_order_id=request.POST.get('order_id')
        

        try:
            
            razorpay_client.utility.verify_payment_signature({
                'razorpay_order_id': order_number,
                'razorpay_payment_id': payment_id,
                'razorpay_signature': signature
            })

            
            order = get_object_or_404(Order, id=db_order_id) 
            order.payment_status = 'Success'
            order.payment_type = 'RazorPay'
            order.save()

            
            if order.payment_status == 'Success':
                for item in order.items.all():
                    item.product.stock -= item.quantity
                    item.product.save()

            
            return redirect('order_success')

        except razorpay.errors.SignatureVerificationError:
            return JsonResponse({'status': 'Payment verification failed'}, safe=False)
        except Exception as e:
            return JsonResponse({'status': str(e)}, safe=False)

    return JsonResponse(order_success)

def order_success(request):
    
    return render(request, 'user/order_success.html')

def process_refund(order_item):
    user = order_item.order.user
    wallet, created = Wallet.objects.get_or_create(user=user, defaults={'balance':Decimal(0.00)})

    order = order_item.order
    original_total = sum(item.subtotal_price for item in order.items.all())
    discount_value = order.total_price - original_total

    item_discount = (Decimal(order_item.subtotal_price) / Decimal(original_total) * Decimal(discount_value)) if original_total > 0 else Decimal('0.00')

    refund_amount = Decimal(order_item.subtotal_price) - item_discount

    WalletTransaction.objects.create(
        wallet=wallet,
        transaction_type='Refund',
        amount=refund_amount,
        created_at=timezone.now()
    )

    wallet.balance += refund_amount
    wallet.save()

@csrf_exempt
def apply_coupon(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            coupon_code = data.get("coupon_code")
            cart_total = Decimal(data.get("cart_total", "0.00"))
            print("Received coupon code:", coupon_code)

            coupon = Coupon.objects.get(code=coupon_code, active=True)

            now = timezone.now().date()


            user_coupon, created = UserCoupon.objects.get_or_create(user=request.user, coupon=coupon)

            if user_coupon.used:
                return JsonResponse({"error": "You have already used this coupon."})
            
            if coupon.coupon_limit <= 0:
                return JsonResponse({"error": "This coupon has reached its usage limit."})

            if coupon.valid_from <= now <= coupon.valid_upto and cart_total >= Decimal(coupon.minimum_purchase_amount):
                request.session['coupon_code'] = coupon.code
                request.session['discount_value'] = str(coupon.discount_value) 
                
                coupon.coupon_limit -= 1
                coupon.save()

                return JsonResponse({"success": True, "discount_value": coupon.discount_value})
            else:
                request.session.pop('coupon_code', None)
                request.session.pop('discount_value', None)
                print("Coupon is not valid for the current date.")
                return JsonResponse({"error": "Coupon is not valid for this purchase."})

        except Coupon.DoesNotExist:
            return JsonResponse({"error": "Invalid coupon code."})
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid request body."})

    return JsonResponse({"error": "Invalid request."})

@csrf_exempt
def remove_coupon(request):
    if request.method == "POST":
        request.session.pop('coupon_code', None)
        request.session.pop('discount_value', None)
        return JsonResponse({"success": True})

    return JsonResponse({"error": "Invalid request."})

def user_return_order_item(request,order_item_id):
    order_item =  get_object_or_404(OrderItem,id=order_item_id)

    if order_item.status == 'Delivered':
        return_reason =  request.POST.get('return_reason','')

        order_item.return_reason =  return_reason
        order_item.status = 'Requested Return'
        order_item.save()
        

    else:
        pass

    return redirect('user_single_order_items',order_id=order_item.order.id)

@login_required(login_url='admin_login')
@never_cache
def sales_report(request):

    startdate = request.GET.get('startdate')
    enddate = request.GET.get('enddate')
    sort_option = request.GET.get('sortOption')

    if startdate:
        try:
            startdate = datetime.strptime(startdate, '%Y-%m-%d').date()
        except ValueError:
            startdate = None
    if enddate:
        try:
            enddate = datetime.strptime(enddate, '%Y-%m-%d').date()
        except ValueError:
            enddate = None
    today = date.today()

    sales_report = (
        Order.objects.annotate(
            date=TruncDate('created_at'),
            total_discount=Coalesce(F('coupon_discount'), Value(0, output_field=DecimalField())) + 
                        Coalesce(F('offer_discount'), Value(0, output_field=DecimalField())),
            net_sales=Sum(
                ExpressionWrapper(
                    F('items__product__price') * F('items__quantity'), 
                    output_field=DecimalField()
                ),
                filter=~Q(items__status__in=['Cancelled', 'Approve Return']),
                output_field=DecimalField()
            )
        )
        .values('date', 'net_sales', 'total_discount', 'coupon_discount', 'offer_discount')  # Include required fields
        .annotate(
            total_sales_revenue=ExpressionWrapper(
                F('net_sales') - F('total_discount'),
                output_field=DecimalField()
            ),
            total_units_sold=Sum(
                'items__quantity',
                filter=~Q(items__status__in=['Cancelled', 'Approve Return']),
                output_field=DecimalField()
            )
        )
        .order_by('date')
    )


    if startdate and enddate:
        sales_report = sales_report.filter(date__range=[startdate, enddate])
    elif startdate:
        sales_report = sales_report.filter(date__gte=startdate)
    elif enddate:
        sales_report = sales_report.filter(date__lte=enddate)


    if sort_option == 'day':
        sales_report = sales_report.filter(date=today)
    elif sort_option == 'week':
        start_of_week = today - timedelta(days=today.weekday())
        sales_report = sales_report.filter(date__gte=start_of_week)
    elif sort_option == 'month':
        sales_report = sales_report.filter(date__month=today.month)
    elif sort_option == 'year':
        sales_report = sales_report.filter(date__year=today.year)

    # Modify total_sales_price calculation to use original prices
    total_sales_price = OrderItem.objects.exclude(
        status__in=['Cancelled', 'Approve Return']
    ).aggregate(
        total=Coalesce(Sum(ExpressionWrapper(
            F('product__price') * F('quantity'), 
            output_field=DecimalField()
        )), Value(0, output_field=DecimalField()))
    )['total']
    
    # Total order count and other calculations remain the same
    total_order_count = Order.objects.count()
    total_units_sold = OrderItem.objects.exclude(
        status__in=['Cancelled', 'Approve Return']
    ).aggregate(
        total=Coalesce(Sum('quantity'), Value(0, output_field=DecimalField()))
    )['total']
    total_order_amount = Order.objects.aggregate(
        total=Coalesce(Sum('total_price'), Value(0, output_field=DecimalField()))
    )['total']

    filtered_order_count = Order.objects.exclude(
        items__status__in=['Cancelled', 'Approve Return']
    ).distinct().count()

    discount_totals = Order.objects.aggregate(
        total_coupon_discount=Coalesce(Sum('coupon_discount'), Value(0, output_field=DecimalField())),
        total_offer_discount=Coalesce(Sum('offer_discount'), Value(0, output_field=DecimalField())),
        total_discount_all=Coalesce(
            Sum(F('coupon_discount') + F('offer_discount')), 
            Value(0, output_field=DecimalField())
        )
    )

    context = {
        'sales_report': sales_report,
        'total_revenue': total_sales_price,
        'total_units_sold': total_units_sold,
        'filtered_order_count': filtered_order_count,
        'total_orders_count': total_order_count,
        'total_order_amount': total_order_amount,
        'total_discount_all': discount_totals['total_discount_all'] or 0,
        'total_coupon_discount': discount_totals['total_coupon_discount'] or 0,
        'total_offer_discount': discount_totals['total_offer_discount'] or 0,
        'startdate': startdate,
        'enddate': enddate,
        'sortOption': sort_option,
    }

    return render(request, 'sales_report.html', context)

def download_sales_report_pdf(request):
    startdate = request.GET.get('startdate')
    enddate = request.GET.get('enddate')
    sort_option = request.GET.get('sortOption')

    if startdate:
        try:
            startdate = datetime.strptime(startdate, '%Y-%m-%d').date()
        except ValueError:
            startdate = None
    if enddate:
        try:
            enddate = datetime.strptime(enddate, '%Y-%m-%d').date()
        except ValueError:
            enddate = None

    today = date.today()

    sales_report = (
        Order.objects.annotate(
            date=TruncDate('created_at'),
            total_discount=Coalesce(F('coupon_discount'), Value(0, output_field=DecimalField())) + 
                        Coalesce(F('offer_discount'), Value(0, output_field=DecimalField())),
            net_sales=Sum(
                ExpressionWrapper(
                    F('items__product__price') * F('items__quantity'), 
                    output_field=DecimalField()
                ),
                filter=~Q(items__status__in=['Cancelled', 'Approve Return']),
                output_field=DecimalField()
            )
        )
        .values('date', 'net_sales', 'total_discount', 'coupon_discount', 'offer_discount')
        .annotate(
            total_sales_revenue=ExpressionWrapper(
                F('net_sales') - F('total_discount'),
                output_field=DecimalField()
            ),
            total_units_sold=Sum(
                'items__quantity',
                filter=~Q(items__status__in=['Cancelled', 'Approve Return']),
                output_field=DecimalField()
            )
        )
        .order_by('date')
    )

    if startdate and enddate:
        sales_report = sales_report.filter(date__range=[startdate, enddate])
    elif startdate:
        sales_report = sales_report.filter(date__gte=startdate)
    elif enddate:
        sales_report = sales_report.filter(date__lte=enddate)

    if sort_option == 'day':
        sales_report = sales_report.filter(date=today)
    elif sort_option == 'week':
        start_of_week = today - timedelta(days=today.weekday())
        sales_report = sales_report.filter(date__gte=start_of_week)
    elif sort_option == 'month':
        sales_report = sales_report.filter(date__month=today.month)
    elif sort_option == 'year':
        sales_report = sales_report.filter(date__year=today.year)

    total_sales_price = OrderItem.objects.exclude(
        status__in=['Cancelled', 'Approve Return']
    ).aggregate(
        total=Coalesce(Sum(ExpressionWrapper(
            F('product__price') * F('quantity'), 
            output_field=DecimalField()
        )), Value(0, output_field=DecimalField()))
    )['total']
    
    total_order_count = Order.objects.count()
    total_units_sold = OrderItem.objects.exclude(
        status__in=['Cancelled', 'Approve Return']
    ).aggregate(
        total=Coalesce(Sum('quantity'), Value(0, output_field=DecimalField()))
    )['total']
    total_order_amount = Order.objects.aggregate(
        total=Coalesce(Sum('total_price'), Value(0, output_field=DecimalField()))
    )['total']

    filtered_order_count = Order.objects.exclude(
        items__status__in=['Cancelled', 'Approve Return']
    ).distinct().count()

    discount_totals = Order.objects.aggregate(
        total_coupon_discount=Coalesce(Sum('coupon_discount'), Value(0, output_field=DecimalField())),
        total_offer_discount=Coalesce(Sum('offer_discount'), Value(0, output_field=DecimalField())),
        total_discount_all=Coalesce(
            Sum(F('coupon_discount') + F('offer_discount')), 
            Value(0, output_field=DecimalField())
        )
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        'Header', fontSize=10, textColor=colors.black, spaceAfter=6, alignment=1
    )
    title_style = ParagraphStyle('Title', fontSize=14, spaceAfter=12, alignment=1)

    company_info = Paragraph(
        "<b>MeatSouk Fresh Market PVT</b><br/>"
        "Email: meatsouk@gmail.com<br/>"
        "Website: www.meatsouk.com", header_style
    )
    elements.append(company_info)
    elements.append(Spacer(1, 12))

    title = f"Sales Report ({startdate or 'All Time'} - {enddate or 'Present'}) - Sorted by {sort_option.capitalize() if sort_option else 'Date'}"
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))

    summary_data = [
        ["Total Revenue", f"{total_sales_price:.2f}"],
        ["Total Orders", total_order_count],
        ["Filtered Orders", filtered_order_count],
        ["Total Units Sold", total_units_sold],
        ["Total Order Amount", f"{total_order_amount:.2f}"],
        ["Total Discount", f"{discount_totals['total_discount_all']:.2f}"],
        ["Coupon Discount", f"{discount_totals['total_coupon_discount']:.2f}"],
        ["Offer Discount", f"{discount_totals['total_offer_discount']:.2f}"]
    ]
    summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 24))

    table_data = [
        ["Date", "Net Sales", "Total Discount", "Coupon Discount", "Offer Discount", "Total Sales Revenue", "Total Units Sold"]
    ]

    for record in sales_report:
        table_data.append([
            record['date'].strftime('%Y-%m-%d') if record.get('date') else 'N/A',
            f"{record.get('net_sales', 0):.2f}",
            f"{record.get('total_discount', 0):.2f}",
            f"{record.get('coupon_discount', 0):.2f}",
            f"{record.get('offer_discount', 0):.2f}",
            f"{record.get('total_sales_revenue', 0):.2f}",
            record.get('total_units_sold', 0)
        ])

    table = Table(table_data, colWidths=[1.5*inch] + [1.2*inch]*6)
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])

    table.setStyle(table_style)
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="MeatSouk_Sales_Report_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response

def download_sales_report_excel(request):
    startdate = request.GET.get('startdate')
    enddate = request.GET.get('enddate')
    sort_option = request.GET.get('sortOption')

    if startdate:
        try:
            startdate = datetime.strptime(startdate, '%Y-%m-%d').date()
        except ValueError:
            startdate = None
    if enddate:
        try:
            enddate = datetime.strptime(enddate, '%Y-%m-%d').date()
        except ValueError:
            enddate = None

    today = date.today()

    sales_report = (
        Order.objects.annotate(
            date=TruncDate('created_at'),
            total_discount=Coalesce(F('coupon_discount'), Value(0, output_field=DecimalField())) + 
                        Coalesce(F('offer_discount'), Value(0, output_field=DecimalField())),
            net_sales=Sum(
                ExpressionWrapper(
                    F('items__product__price') * F('items__quantity'), 
                    output_field=DecimalField()
                ),
                filter=~Q(items__status__in=['Cancelled', 'Approve Return']),
                output_field=DecimalField()
            )
        )
        .values('date', 'net_sales', 'total_discount', 'coupon_discount', 'offer_discount')
        .annotate(
            total_sales_revenue=ExpressionWrapper(
                F('net_sales') - F('total_discount'),
                output_field=DecimalField()
            ),
            total_units_sold=Sum(
                'items__quantity',
                filter=~Q(items__status__in=['Cancelled', 'Approve Return']),
                output_field=DecimalField()
            )
        )
        .order_by('date')
    )

    if startdate and enddate:
        sales_report = sales_report.filter(date__range=[startdate, enddate])
    elif startdate:
        sales_report = sales_report.filter(date__gte=startdate)
    elif enddate:
        sales_report = sales_report.filter(date__lte=enddate)

    if sort_option == 'day':
        sales_report = sales_report.filter(date=today)
    elif sort_option == 'week':
        start_of_week = today - timedelta(days=today.weekday())
        sales_report = sales_report.filter(date__gte=start_of_week)
    elif sort_option == 'month':
        sales_report = sales_report.filter(date__month=today.month)
    elif sort_option == 'year':
        sales_report = sales_report.filter(date__year=today.year)

    # Total calculations matching sales report view
    total_sales_price = OrderItem.objects.exclude(
        status__in=['Cancelled', 'Approve Return']
    ).aggregate(
        total=Coalesce(Sum(ExpressionWrapper(
            F('product__price') * F('quantity'), 
            output_field=DecimalField()
        )), Value(0, output_field=DecimalField()))
    )['total']
    
    total_order_count = Order.objects.count()
    total_units_sold = OrderItem.objects.exclude(
        status__in=['Cancelled', 'Approve Return']
    ).aggregate(
        total=Coalesce(Sum('quantity'), Value(0, output_field=DecimalField()))
    )['total']
    total_order_amount = Order.objects.aggregate(
        total=Coalesce(Sum('total_price'), Value(0, output_field=DecimalField()))
    )['total']

    filtered_order_count = Order.objects.exclude(
        items__status__in=['Cancelled', 'Approve Return']
    ).distinct().count()

    discount_totals = Order.objects.aggregate(
        total_coupon_discount=Coalesce(Sum('coupon_discount'), Value(0, output_field=DecimalField())),
        total_offer_discount=Coalesce(Sum('offer_discount'), Value(0, output_field=DecimalField())),
        total_discount_all=Coalesce(
            Sum(F('coupon_discount') + F('offer_discount')), 
            Value(0, output_field=DecimalField())
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    ws['A1'] = "Sales Report Summary"
    ws['A1'].font = Font(bold=True, size=14)
    
    summary_data = [
        ("Total Revenue", f"{total_sales_price:.2f}"),
        ("Total Orders", total_order_count),
        ("Filtered Orders", filtered_order_count),
        ("Total Units Sold", total_units_sold),
        ("Total Order Amount", f"{total_order_amount:.2f}"),
        ("Total Discount", f"{discount_totals['total_discount_all']:.2f}"),
        ("Coupon Discount", f"{discount_totals['total_coupon_discount']:.2f}"),
        ("Offer Discount", f"{discount_totals['total_offer_discount']:.2f}")
    ]

    for row_num, (label, value) in enumerate(summary_data, start=3):
        ws[f'A{row_num}'] = label
        ws[f'B{row_num}'] = value
        ws[f'A{row_num}'].font = Font(bold=True)
    
    headers = ["Date", "Net Sales", "Total Discount", "Coupon Discount", "Offer Discount", "Total Sales Revenue", "Total Units Sold"]
    ws.append(headers)
    for cell in ws[len(summary_data) + 3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for record in sales_report:
        ws.append([
            record['date'].strftime('%Y-%m-%d') if record.get('date') else 'N/A',
            record.get('net_sales', 0),
            record.get('total_discount', 0),
            record.get('coupon_discount', 0),
            record.get('offer_discount', 0),
            record.get('total_sales_revenue', 0),
            record.get('total_units_sold', 0)
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="MeatSouk_Sales_Report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

def generate_invoice_number():
    """Generate a unique invoice number with pattern 'meat123123XXX'."""
    unique_id = random.randint(100000, 999999)
    return f"meat{unique_id}"
def render_to_pdf(template_src, context_dict):
    template = render_to_string(template_src, context_dict)
    result = BytesIO()
    pdf = pisa.CreatePDF(template, dest=result)
    if not pdf.err:
        return result.getvalue()
    return None
def download_invoice_item(request, item_id):
    order_item = get_object_or_404(OrderItem, id=item_id, order__user=request.user)
    order = order_item.order

    if order_item.status != "Delivered":
        return HttpResponse("Invoice can only be downloaded for delivered items", status=400)

    invoice, created = Invoice.objects.get_or_create(
        order_item=order_item,
        defaults={
            'invoice_number': generate_invoice_number(),
            'invoice_date': timezone.now()
        }
    )

    if not created:
        invoice.invoice_date = timezone.now()
        invoice.save()

    coupon_discount = order.coupon_discount
    offer_discount = order.offer_discount

    context = {
        'order_item': order_item,
        'invoice_number': invoice.invoice_number,
        'invoice_date': invoice.invoice_date,
        'order': order,
        'coupon_discount': coupon_discount,
        'offer_discount': offer_discount,
        'total_discount': coupon_discount + offer_discount,
        'subtotal': order_item.price,
    }

    pdf = render_to_pdf('user/order_invoice.html', context)
    
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="invoice_{order_item.id}.pdf"'
        return response
    
    return HttpResponse("PDF generation failed", status=500)